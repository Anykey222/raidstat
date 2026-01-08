from abc import ABC, abstractmethod
import pandas as pd
import os
from datetime import datetime
import logging
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
import io

class StorageInterface(ABC):
    @abstractmethod
    def get_roster(self, source="statistics"):
        """Возвращает список известных имен из указанного источника."""
        pass

    @abstractmethod
    def save_attendance(self, attendance_data, date_str):
        """Сохраняет данные о посещаемости."""
        pass

    @abstractmethod
    def save_statistics(self, stats_data, date_str):
        """Сохраняет данные статистики."""
        pass

class ExcelStorage(StorageInterface):
    def __init__(self, file_path="Raidstat.xlsx"):
        self.file_path = file_path
        self.logger = logging.getLogger(__name__)
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        if not os.path.exists(self.file_path):
            # Создаем пустой excel с необходимыми листами и заголовками по умолчанию
            wb = openpyxl.Workbook()
            
            # Лист Посещаемость
            ws_att = wb.create_sheet("Посещаемость")
            ws_att.cell(row=1, column=1, value="Ник")
            
            # Лист Статистика
            ws_stat = wb.create_sheet("Статистика")
            fixed_headers = ["Хонор до", "Хонор после", "Фраги до", "Фраги после", "ГС", "Ник", "Класс"]
            for i, header in enumerate(fixed_headers, start=1):
                ws_stat.cell(row=1, column=i, value=header)
                
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            wb.save(self.file_path)
            self.logger.info(f"Создан новый файл {self.file_path} с заголовками по умолчанию")

    def get_roster(self, source="statistics"):
        names = set()
        try:
            if not os.path.exists(self.file_path):
                return []

            # Используем ExcelFile для оптимизации проверки листов с контекстным менеджером
            with pd.ExcelFile(self.file_path) as xls:
                # 1. Имена из статистики
                if source == "statistics" and "Статистика" in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name="Статистика")
                    if not df.empty and 'Ник' in df.columns:
                        # Нормализуем имена: убираем пробелы и приводим к строке
                        current_names = df['Ник'].dropna().astype(str).str.strip().tolist()
                        names.update(current_names)
                
                # 2. Имена из посещаемости
                elif source == "attendance" and "Посещаемость" in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name="Посещаемость")
                    if not df.empty:
                        # Пытаемся найти колонку 'Ник', иначе используем первую
                        col_name = 'Ник' if 'Ник' in df.columns else df.columns[0]
                        current_names = df[col_name].dropna().astype(str).str.strip().tolist()
                        # Отфильтровываем возможные синонимы заголовка или пустые строки, если определение заголовка не удалось
                        names.update([n for n in current_names if n.lower() != 'ник'])

            # Удаляем пустые строки
            names.discard("")
            return list(names)

        except Exception as e:
            self.logger.error(f"Ошибка при чтении ростера: {e}")
            return []

    def save_attendance(self, attendance_data, date_str):
        """
        attendance_data: Список присутствующих имен.
        """
        try:
            # Используем openpyxl напрямую для добавления колонок без перезаписи/потери форматирования
            if not os.path.exists(self.file_path):
                self._ensure_file_exists()

            wb = openpyxl.load_workbook(self.file_path)
            
            if "Посещаемость" not in wb.sheetnames:
                ws = wb.create_sheet("Посещаемость")
            else:
                ws = wb["Посещаемость"]

            # Привязываем существующие имена к строкам (колонка 1)
            name_to_row = {}
            max_row = ws.max_row
            
            # Проверяем, существует ли заголовок "Ник"
            if ws.cell(row=1, column=1).value != "Ник":
                ws.cell(row=1, column=1, value="Ник")

            # Собираем существующие имена
            for row in range(2, max_row + 1):
                cell_val = ws.cell(row=row, column=1).value
                if cell_val:
                    name_to_row[str(cell_val).strip()] = row
            
            # Находим следующую свободную колонку
            # max_column — последняя колонка с данными.
            next_col = ws.max_column + 1
            
            # Записываем заголовок даты
            ws.cell(row=1, column=next_col, value=str(date_str))
            
            # Инициализируем 0 для всех существующих участников
            for row_idx in name_to_row.values():
                 ws.cell(row=row_idx, column=next_col, value=0)
            
            # Записываем посещаемость
            for name in attendance_data:
                name = str(name).strip()
                if not name:
                    continue
                    
                if name in name_to_row:
                    row_idx = name_to_row[name]
                    ws.cell(row=row_idx, column=next_col, value=1)
                else:
                    # Добавляем новое имя в конец
                    new_row = ws.max_row + 1
                    # Двойная проверка, чтобы не перезаписать (логика max_row обрабатывает это)
                    ws.cell(row=new_row, column=1, value=name)
                    ws.cell(row=new_row, column=next_col, value=1)
                    name_to_row[name] = new_row
            
            wb.save(self.file_path)
            self.logger.info(f"Сохранена колонка посещаемости: {date_str} (Всего участников: {len(attendance_data)})")
                
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении посещаемости: {e}")
            import traceback
            self.logger.error(traceback.format_exc())

    def save_statistics(self, stats_data, date_str, debug_screens=False):
        """
        stats_data: Словарь {имя: {kills: int, honor: int, gear: int, class: str, ...}}
        Обновляет лист 'Статистика'.
        Колонки: ГС, Ник, Класс, ... [Хонор {дата}, Фраги {дата}, Очки]
        В колонках события сохраняются дельты (разница между концом и началом).
        
        Если debug_screens=True, добавляет отладочные изображения как комментарии к ячейкам.
        """
        try:
            if not os.path.exists(self.file_path):
                self._ensure_file_exists()
            
            # Загружаем с data_only=True, чтобы "заморозить" значения из предыдущих запусков.
            # Это критично, так как колонки 1-7 (Начало/Конец) общие/перезаписываются для каждого события.
            # Только НОВОЕ событие должно иметь активные формулы, указывающие на эти общие колонки.
            # Старые события должны быть заморожены до статических значений для сохранения их истории.
            wb = openpyxl.load_workbook(self.file_path, data_only=True)

            if "Статистика" not in wb.sheetnames:
                ws = wb.create_sheet("Статистика")
            else:
                ws = wb["Статистика"]

            # Удаляем все комментарии в первых 7 колонках, чтобы избавиться от 
            # "висячих" подсказок с пустыми областями от предыдущих запусков
            for row in ws.iter_rows(min_col=1, max_col=7):
                for cell in row:
                    cell.comment = None

            # Фиксированные заголовки: Хонор до, Хонор после, Фраги до, Фраги после, ГС, Ник, Класс
            fixed_headers = ["Хонор до", "Хонор после", "Фраги до", "Фраги после", "ГС", "Ник", "Класс"]
            
            # Проверяем, есть ли заголовки. Если нет - записываем.
            first_cell = ws.cell(row=1, column=1).value
            if not first_cell or first_cell != fixed_headers[0]:
                for i, header in enumerate(fixed_headers, start=1):
                    ws.cell(row=1, column=i, value=header)

            # Определяем начальную колонку для нового события
            # Она должна быть после последней колонки.
            start_col = ws.max_column + 1
            
            # Добавляем заголовки нового события
            # "Хонор " + дата, "Фраги " + дата, "Очки"
            new_headers = [f"Хонор {date_str}", f"Фраги {date_str}", "Очки"]
            
            ws.cell(row=1, column=start_col, value=new_headers[0])
            ws.cell(row=1, column=start_col+1, value=new_headers[1])
            ws.cell(row=1, column=start_col+2, value=new_headers[2])

            # Мапим имена на номера строк
            name_to_row = {}
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ник в колонке 6 (индекс 5)
                if len(row) > 5:
                    name = row[5]  # Колонка 6 (Ник)
                    if name:
                        name_to_row[str(name)] = i

            from openpyxl.utils import get_column_letter

            # Обновляем/добавляем данные
            for name, data in stats_data.items():
                row_idx = name_to_row.get(name)
                
                if not row_idx:
                    # Добавляем новую строку
                    row_idx = ws.max_row + 1
                    ws.cell(row=row_idx, column=6, value=name)  # Ник в колонке 6
                
                # Обновляем статические колонки (ПЕРЕЗАПИСЫВАЮТСЯ при каждом вызове)
                # Колонка 1: Хонор до (A)
                val = data.get('honor_start')
                ws.cell(row=row_idx, column=1, value=val if val is not None else "")
                # Колонка 2: Хонор после (B)
                val = data.get('honor_end')
                ws.cell(row=row_idx, column=2, value=val if val is not None else "")
                # Колонка 3: Фраги до (C)
                val = data.get('kills_start')
                ws.cell(row=row_idx, column=3, value=val if val is not None else "")
                # Колонка 4: Фраги после (D)
                val = data.get('kills_end')
                ws.cell(row=row_idx, column=4, value=val if val is not None else "")
                
                # Колонка 5: ГС (E)
                val = data.get('gear')
                ws.cell(row=row_idx, column=5, value=val if val is not None else "")
                # Колонка 6: Ник (F) (уже установлен выше)
                # Колонка 7: Класс (G)
                ws.cell(row=row_idx, column=7, value=data.get('class', ''))
                
                # Обновляем колонки события (дельты!) - ДОБАВЛЯЮТСЯ В КОНЕЦ
                # Вместо статических значений пишем формулы для автоматического пересчета
                
                # Координаты ячеек для формулы
                # A=1, B=2, C=3, D=4
                col_hon_start = "A" # Хонор до
                col_hon_end = "B"   # Хонор после
                col_kills_start = "C" # Фраги до
                col_kills_end = "D"   # Фраги после
                
                # Колонки для результата
                col_res_honor = get_column_letter(start_col)
                col_res_kills = get_column_letter(start_col+1)
                # col_res_points = get_column_letter(start_col+2)

                # Формула Хонор Дельта = B - A
                # Используем IF и ISNUMBER для корректной обработки пустых ячеек ("")
                formula_honor = f'=IF(AND(ISNUMBER({col_hon_start}{row_idx}), ISNUMBER({col_hon_end}{row_idx})), {col_hon_end}{row_idx}-{col_hon_start}{row_idx}, "")'
                ws.cell(row=row_idx, column=start_col, value=formula_honor)
                
                # Формула Фраги Дельта = D - C
                formula_kills = f'=IF(AND(ISNUMBER({col_kills_start}{row_idx}), ISNUMBER({col_kills_end}{row_idx})), {col_kills_end}{row_idx}-{col_kills_start}{row_idx}, "")'
                ws.cell(row=row_idx, column=start_col+1, value=formula_kills)
                
                # Формула Очки = Фраги * 70 + Хонор
                # Ссылаемся на только что созданные ячейки дельт
                formula_points = f'=IF(AND(ISNUMBER({col_res_kills}{row_idx}), ISNUMBER({col_res_honor}{row_idx})), {col_res_kills}{row_idx}*70+{col_res_honor}{row_idx}, "")'
                ws.cell(row=row_idx, column=start_col+2, value=formula_points)
                
                # Добавляем отладочные изображения как комментарии (если debug_screens=True)
                if debug_screens:
                    # Проверяем наличие скриншотов (start или end)
                    if data.get('debug_images_start') or data.get('debug_images_end'):
                        self._add_debug_images_comment(ws, row_idx, data, name)

            # Применяем автофильтр и сортировку
            self._apply_autofilter_and_sort(ws, start_col + 2)  # Колонка "Очки"

            # Получаем список изображений для комментариев до сохранения
            comment_images = getattr(wb, '_debug_comment_images', None)
            
            wb.save(self.file_path)
            
            # После сохранения openpyxl — применяем сортировку и изображения через COM
            # (COM нужен для корректной сортировки по формулам и вставки изображений)
            self._apply_com_operations(comment_images, start_col + 2)  # start_col+2 = колонка "Очки"

        except Exception as e:
            self.logger.error(f"Ошибка при сохранении статистики: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
    
    def _apply_autofilter_and_sort(self, ws, points_col):
        """
        Применяет автофильтр ко всем колонкам.
        Сортировка выполняется через COM после сохранения файла.
        
        points_col: номер колонки с полем "Очки" (1-indexed) - не используется здесь
        """
        from openpyxl.utils import get_column_letter
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row < 2:
            return  # Нет данных
        
        # Диапазон для автофильтра: A1 до последней колонки/строки
        filter_range = f"A1:{get_column_letter(max_col)}{max_row}"
        
        # Устанавливаем автофильтр
        ws.auto_filter.ref = filter_range
    
    def _add_debug_images_comment(self, ws, row_idx, data, name):
        """
        Сохраняет информацию для вставки изображений в комментарии ко всем 7 столбцам:
        Хонор до, Хонор после, Фраги до, Фраги после, ГС, Ник, Класс
        
        data должен содержать debug_images_start и debug_images_end
        """
        try:
            wb = ws.parent
            
            if not hasattr(wb, '_debug_comment_images'):
                wb._debug_comment_images = []
            
            # Получаем скриншоты для "до" и "после"
            images_start = data.get('debug_images_start') or {}
            images_end = data.get('debug_images_end') or {}
            
            # Маппинг колонок к изображениям
            # Колонка 1 (Хонор до) -> honor из start
            # Колонка 2 (Хонор после) -> honor из end
            # Колонка 3 (Фраги до) -> kills из start
            # Колонка 4 (Фраги после) -> kills из end
            # Колонка 5 (ГС) -> gear из end
            # Колонка 6 (Ник) -> name из end
            # Колонка 7 (Класс) -> class из end
            column_mappings = [
                (1, images_start, 'honor'),   # Хонор до
                (2, images_end, 'honor'),     # Хонор после
                (3, images_start, 'kills'),   # Фраги до
                (4, images_end, 'kills'),     # Фраги после
                (5, images_end, 'gear'),      # ГС
                (6, images_end, 'name'),      # Ник
                (7, images_end, 'class')      # Класс
            ]
            
            for col, images_dict, image_field in column_mappings:
                img_path = images_dict.get(image_field) if images_dict else None
                
                # Fallback для ГС, Ника и Класса (колонки 5, 6, 7):
                # Если нет изображения "после" (images_end), пробуем взять из "до" (images_start).
                # Это актуально для пропущенных участников (перенесённых с пред. группы),
                # у которых нет "end", но есть "start" (как "end" предыдущей группы).
                if not img_path and col in [5, 6, 7]:
                    img_path = images_start.get(image_field) if images_start else None

                wb._debug_comment_images.append({
                    'sheet': ws.title,
                    'row': row_idx,
                    'name': name, # Store name for mapping after sort
                    'col': col,
                    'image_path': img_path if img_path and os.path.exists(img_path) else None
                })
            
        except Exception as e:
            self.logger.warning(f"Ошибка добавления отладочных изображений: {e}")
    
    def add_images_to_comments(self, comment_images):
        """
        Добавляет изображения в комментарии Excel через COM-интерфейс (pywin32).
        
        Эта функция должна быть вызвана ПОСЛЕ сохранения файла openpyxl.
        Использует Excel COM API для добавления изображений в фон комментариев.
        
        - Если у ячейки есть комментарий — он перезаписывается
        - Если нет изображения — комментарий удаляется
        
        comment_images: список словарей с ключами:
            - sheet: название листа
            - row: номер строки (1-indexed)
            - col: номер колонки (1-indexed)  
            - image_path: путь к изображению (или None для удаления)
        """
        if not comment_images:
            return
        
        try:
            import win32com.client
            import pythoncom
            
            # Инициализируем COM
            pythoncom.CoInitialize()
            
            added_count = 0
            removed_count = 0
            
            excel = None
            wb = None
            try:
                # Открываем Excel
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                
                # Открываем файл (используем абсолютный путь)
                abs_path = os.path.abspath(self.file_path)
                wb = excel.Workbooks.Open(abs_path)
                
                try:
                    for img_info in comment_images:
                        try:
                            sheet_name = img_info['sheet']
                            row = img_info['row']
                            col = img_info['col']
                            image_path = img_info['image_path']
                            
                            # Получаем лист и ячейку
                            ws = wb.Worksheets(sheet_name)
                            cell = ws.Cells(row, col)
                            
                            # Если нет изображения — удаляем комментарий
                            if not image_path or not os.path.exists(image_path):
                                if cell.Comment is not None:
                                    cell.Comment.Delete()
                                    removed_count += 1
                                continue
                            
                            # Удаляем существующий комментарий если есть
                            if cell.Comment is not None:
                                cell.Comment.Delete()
                            
                            # Создаём новый комментарий
                            cell.AddComment("")
                            comment = cell.Comment
                            
                            # Получаем Shape комментария
                            shape = comment.Shape
                            
                            # Устанавливаем изображение как фон комментария
                            fill = shape.Fill
                            fill.UserPicture(os.path.abspath(image_path))
                            
                            # Настраиваем размеры комментария под изображение
                            from PIL import Image as PILImage
                            with PILImage.open(image_path) as img:
                                width, height = img.size
                                # Масштабируем для удобного просмотра
                                scale = 2.0
                                shape.Width = width * scale
                                shape.Height = height * scale
                            
                            added_count += 1
                            
                        except Exception as e:
                            self.logger.warning(f"Ошибка добавления изображения в комментарий [{row},{col}]: {e}")
                    
                    self.logger.info(f"Комментарии: добавлено {added_count}, удалено {removed_count}")
                    
                    # Сохраняем и закрываем
                    wb.Save()
                    
                finally:
                    if wb:
                        wb.Close(False)
                    
            finally:
                if excel:
                    excel.Quit()
                pythoncom.CoUninitialize()
            
        except ImportError:
            self.logger.warning("pywin32 не установлен. Изображения в комментариях недоступны.")
        except Exception as e:
            # Ошибка -2147221005 (Invalid class string) означает, что Excel не установлен
            if "Invalid class string" in str(e):
                self.logger.warning("Excel не установлен в системе. Пропуск вставки изображений.")
            else:
                self.logger.error(f"Ошибка добавления изображений в комментарии: {e}")
                import traceback
                self.logger.error(traceback.format_exc())

    def _apply_com_operations(self, comment_images, points_col):
        """
        Применяет операции через COM-интерфейс Excel:
        1. Скрытие старых колонок статистики
        2. Сортировка по колонке "Очки" (Специальная логика: числа убывают, пустые строки внизу)
        3. Добавление изображений в комментарии (используя привязку по имени)
        
        Эта функция должна быть вызвана ПОСЛЕ сохранения файла openpyxl.
        """
        try:
            import win32com.client
            import pythoncom
            from openpyxl.formula.translate import Translator
            
            # Инициализация COM
            pythoncom.CoInitialize()
            
            excel = None
            wb = None
            try:
                # Открываем Excel
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                
                # Открываем файл
                abs_path = os.path.abspath(self.file_path)
                wb = excel.Workbooks.Open(abs_path)
                
                try:
                    ws = wb.Worksheets("Статистика")
                    
                    # === 1. СКРЫТИЕ СТАРЫХ КОЛОНОК ===
                    hide_start = 8
                    hide_end = points_col - 3
                    
                    if hide_end >= hide_start:
                        col_start_char = self._get_col_letter(hide_start)
                        col_end_char = self._get_col_letter(hide_end)
                        range_str = f"{col_start_char}:{col_end_char}"
                        ws.Columns(range_str).Hidden = True
                        self.logger.info(f"Скрыты колонки {range_str}")

                    # === 2. СОРТИРОВКА (Custom Python Sort) ===
                    # Находим используемый диапазон
                    last_row = ws.Cells(ws.Rows.Count, "A").End(-4162).Row # xlUp
                    last_col = ws.UsedRange.Columns.Count
                    
                    name_map = {} # Маппинг Имя -> Новый индекс строки
                    
                    if last_row >= 2:
                        # Читаем данные (Values) для ключа сортировки
                        # Читаем формулы (Formula) для перезаписи
                        # Диапазон: Строка 2...LastRow, Колонка 1...LastCol
                        
                        # индексы COM range начинаются с 1
                        rng = ws.Range(ws.Cells(2, 1), ws.Cells(last_row, last_col))
                        
                        # Если всего одна строка данных, range.Value возвращает не tuple, а single value?
                        # Проверка:
                        if last_row == 2:
                            # Одиночная строка - преобразуем в список списков
                            vals = (rng.Value, ) 
                            formulas = (rng.Formula, )
                            # rng.Value может вернуть tuple, если колонок много, или scalar, если 1 ячейка?
                            # Но колонок точно много (минимум 7). Значит вернет tuple (1D).
                            # Нам нужна консистентная 2D структура.
                            # Если 1 строка, win32com возвращает tuple
                            # Если много - tuple of tuples
                            # Проверим тип
                            if not isinstance(vals[0], tuple):
                                vals = (vals,)
                                formulas = (formulas,)
                        else:
                            vals = rng.Value
                            formulas = rng.Formula
                            
                        # Создаем список строк для сортировки
                        rows = []
                        for i in range(len(vals)):
                            rows.append({
                                'val': vals[i],        # Кортеж вычисленных значений
                                'formula': formulas[i] # Кортеж формул/значений
                            })
                            
                        # Ключ сортировки: данные в колонке points_col
                        # Индекс в кортеже = points_col - 1
                        p_idx = points_col - 1
                        
                        def sort_key(row_dict):
                            val_tuple = row_dict['val']
                            if p_idx < len(val_tuple):
                                v = val_tuple[p_idx]
                                # Группируем: Числа (1) и Остальное (0)
                                if isinstance(v, (int, float)):
                                    return (1, v)
                            return (0, 0) # Не числа (None, "", строки) попадают в группу 0
                        
                        # Сортировка: Группа 1 (числа) выше группы 0. Внутри группы — по убыванию.
                        rows.sort(key=sort_key, reverse=True)
                        
                        # Подготовка данных для записи
                        # Нам нужно обновить относительные ссылки в формулах!
                        new_formulas = []
                        
                        # Мы потеряли исходный индекс при сортировке. Нужно было сохранить!
                        # Исправим алгоритм сбора rows:
                        
                        # Пересоберем rows с индексом
                        rows_indexed = []
                        for i in range(len(vals)):
                            rows_indexed.append({
                                'val': vals[i],
                                'formula': formulas[i],
                                'orig_row': i + 2
                            })
                        
                        rows_indexed.sort(key=sort_key, reverse=True)
                        
                        # Теперь транслируем формулы
                        final_data = [] # Список кортежей для обратной записи
                        
                        for i, item in enumerate(rows_indexed):
                            # Текущая строка назначения (i + 2)
                            new_row = i + 2
                            # Исходная строка (мы не знаем её номер, но знаем формулу)
                            # Проблема: чтобы перевести формулу, нам нужен origin.
                            # Но мы взяли формулу как строку.
                            # Если формула была в Row X, и ссылалась на A{X}, то строка формулы: "=A{X}..."
                            # Если мы пишем её в Row Y, она останется "=A{X}...", если не переведем.
                            
                            # Решение: перед чтением формул мы не знали их Row. 
                            # Но мы читали последовательно с 2 строки. 
                            # Значит formula[k] была из Row k+2.
                            
                            # Мы потеряли исходный индекс при сортировке. Нужно было сохранить!
                            # Исправим алгоритм сбора rows:
                            pass
                        
                        # Пересоберем rows с индексом
                        rows_indexed = []
                        for i in range(len(vals)):
                            rows_indexed.append({
                                'val': vals[i],
                                'formula': formulas[i],
                                'orig_row': i + 2
                            })
                        
                        rows_indexed.sort(key=sort_key, reverse=True)
                        
                        # Теперь транслируем формулы
                        final_data = [] # Список кортежей для обратной записи
                        
                        for i, item in enumerate(rows_indexed):
                            new_row = i + 2
                            orig_row = item['orig_row']
                            
                            # Собираем данные строки
                            row_content = []
                            src_formulas = item['formula']
                            
                            # Мапим Имя -> Новая строка (Колонка 6 — Ник)
                            # Индекс 5
                            # Значение берем из vals (вычисленного), так надежнее, чем из формулы
                            n_val = item['val'][5]
                            if n_val:
                                name_map[str(n_val)] = new_row
                            
                            if new_row == orig_row:
                                # Оптимизация: если строка осталась на месте, формулы не меняем
                                final_data.append(src_formulas)
                            else:
                                # Строка переместилась — транслируем формулы
                                translated_row = []
                                for col_idx, cell_val in enumerate(src_formulas):
                                    # col_idx имеет индекс 0
                                    cell_col_letter = self._get_col_letter(col_idx + 1)
                                    
                                    if isinstance(cell_val, str) and cell_val.startswith('='):
                                        try:
                                            # Трансляция
                                            # Источник: {Col}{OrigRow}
                                            # Назначение: {Col}{NewRow}
                                            origin_addr = f"{cell_col_letter}{orig_row}"
                                            dest_addr = f"{cell_col_letter}{new_row}"
                                            
                                            new_form = Translator(cell_val, origin=origin_addr).translate_formula(dest_addr)
                                            translated_row.append(new_form)
                                        except Exception:
                                            # Резервный вариант
                                            translated_row.append(cell_val)
                                    else:
                                        translated_row.append(cell_val)
                                final_data.append(tuple(translated_row))
                        
                        # Записываем обратно (массивом для скорости)
                        # Преобразуем list of tuples -> tuple of tuples
                        # COM требует Tuple of Tuples
                        write_data = tuple(final_data)
                        rng.Formula = write_data
                        
                        self.logger.info(f"Данные отсортированы (Custom Python Sort).")
                    
                    # === 3. ИЗОБРАЖЕНИЯ В КОММЕНТАРИЯХ ===
                    if comment_images:
                        added_count = 0
                        
                        # Удаляем ВСЕ старые комментарии в обрабатываемом диапазоне? 
                        # Или точечно?
                        # Проще удалять точечно перед добавлением, но мы не знаем, где старые, если они переместились.
                        # Но комментарии ПРИВЯЗАНЫ к ячейкам. При записи Range.Formula комментарии НЕ перемещаются.
                        # Значит, комментарии остались на старых местах (которые теперь занимают другие люди).
                        # Это ПЛОХО. Нам нужно очистить комментарии во всем диапазоне данных перед добавлением новых.
                        # Диапазон данных (колонки 1-7 + новые колонки).
                        # Но у нас есть комментарии только в колонках 1-7.
                        
                        # Очистка комментариев в статических колонках (1-7)
                        clear_rng = ws.Range(ws.Cells(2, 1), ws.Cells(last_row, 7))
                        clear_rng.ClearComments()
                        
                        for img_info in comment_images:
                            try:
                                name = img_info.get('name')
                                # Находим новую строку по имени
                                row = name_map.get(name)
                                
                                if not row:
                                    continue
                                    
                                col = img_info['col']
                                image_path = img_info['image_path']
                                
                                cell = ws.Cells(row, col)
                                
                                if not image_path or not os.path.exists(image_path):
                                    continue
                                
                                cell.AddComment("")
                                comment = cell.Comment
                                shape = comment.Shape
                                
                                fill = shape.Fill
                                fill.UserPicture(os.path.abspath(image_path))
                                
                                from PIL import Image as PILImage
                                with PILImage.open(image_path) as img:
                                    width, height = img.size
                                    scale = 2.0
                                    shape.Width = width * scale
                                    shape.Height = height * scale
                                
                                added_count += 1
                                
                            except Exception as e:
                                self.logger.warning(f"Ошибка добавления изображения: {e}")
                        
                        self.logger.info(f"Комментарии обновлены: {added_count}")
                    
                    wb.Save()
                    
                finally:
                    if wb:
                        wb.Close(False)
                    
            finally:
                if excel:
                    excel.Quit()
                pythoncom.CoUninitialize()
            
        except ImportError:
            self.logger.warning("pywin32 не установлен.")
        except Exception as e:
            # Ошибка -2147221005 (Invalid class string) означает, что Excel не установлен
            if "Invalid class string" in str(e):
                self.logger.warning("Excel не установлен в системе. Пропуск COM-операций (сортировка и оформление).")
            else:
                self.logger.error(f"Ошибка COM-операций: {e}")
                import traceback
                self.logger.error(traceback.format_exc())

    def _get_col_letter(self, col_idx):
        """Helper to get column letter from index (1->A, 2->B, etc)"""
        string = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            string = chr(65 + remainder) + string
        return string

